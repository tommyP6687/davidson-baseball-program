from flask import Flask, render_template, request, send_file
import pandas as pd
import numpy as np
import io
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

app = Flask(__name__)

#Utility Functions
def load_and_filter_trackman(file_objs):
    """
    Load and filter Trackman CSV files.

    Parameters:
    file_objs (list): A list of uploaded Trackman file objects

    Returns:
    pd.DataFrame: Combined and cleaned dataframe of all Trackman data
    """
    desired_columns = ["Batter", "PlateLocHeight", "PlateLocSide", "PitchCall", "KorBB", "PlayResult"]
    filtered_dfs = []

    for f in file_objs:
        df = pd.read_csv(f)
        df.columns = df.columns.str.strip()
        filtered = df[[col for col in desired_columns if col in df.columns]].copy()
        filtered = filtered.dropna(subset=["Batter"])
        
        #Skip empty or all-NA DataFrames to avoid future pandas warning
        if not filtered.empty and not filtered.isna().all().all():
            filtered_dfs.append(filtered)

    # If no valid DataFrames, return an empty DataFrame with correct columns
    if not filtered_dfs:
        return pd.DataFrame(columns=desired_columns)
    
    return pd.concat(filtered_dfs, ignore_index=True)

def calculate_batter_scores(df):
    """
    Compute decision scores for each batter.

    Parameters:
    df (pd.DataFrame): Cleaned Trackman dataframe

    Returns:
    dict: Mapping of batter names to their scores
    """
    scores = {}

    for _, row in df.iterrows():
        batter = row['Batter']
        pitch_call = row['PitchCall']
        korbb = row.get('KorBB', "")
        play_result = row.get('PlayResult', "")
        height = row.get('PlateLocHeight', None)
        side = row.get('PlateLocSide', None)

        if pd.isna(batter):
            continue
        if batter not in scores:
            scores[batter] = 0
            
        #Scoring logic based on pitch call and location
        if pitch_call == "BallCalled":
            scores[batter] += 0.25
        elif pitch_call == "StrikeCalled" and korbb == "Strikeout":
            scores[batter] -= 2
        elif pitch_call == "StrikeSwinging":
            if korbb == "Strikeout":
                scores[batter] -= 1.5
            elif pd.notna(height) and pd.notna(side):
                if not (1.5 <= height <= 3.6 and -0.75 <= side <= 0.75):
                    scores[batter] -= 1
        if play_result == "HomeRun":
            scores[batter] += 4

    return scores

def get_grade(score):
    """
    Convert numerical score into a letter grade.

    Parameters:
    score (float): Decision score

    Returns:
    str: Letter grade
    """
    if score < -0.5:
        return "C-"
    elif score < 0.0:
        return "C"
    elif score < 0.5:
        return "C+"
    elif score < 1.0:
        return "B-"
    elif score < 1.5:
        return "B"
    elif score < 2.0:
        return "B+"
    elif score < 2.5:
        return "A-"
    elif score < 3.0:
        return "A"
    else:
        return "A+"

#Routes
@app.route('/')
def index():
    """Render the homepage with upload form."""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    """Handle file uploads, compute scores, merge data, and return Excel output."""
    try:
        trackman_files = request.files.getlist('trackman')
        trumedia_file = request.files.get('trumedia')

        #1: Load Trackman and compute scores
        combined_df = load_and_filter_trackman(trackman_files)
        scores = calculate_batter_scores(combined_df)
        
        #2: Create score dataframe with grades and normalized names
        score_df = pd.DataFrame(scores.items(), columns=["Batter", "decisionScore"])
        score_df["Grade"] = score_df["decisionScore"].apply(get_grade)
        score_df["NormalizedName"] = score_df["Batter"].apply(
            lambda b: ' '.join(b.replace('"', '').replace(',', '').split()[::-1])
        )
        score_map = {
            name: {"Grade": grade, "decisionScore": score}
            for name, grade, score in zip(score_df["NormalizedName"], score_df["Grade"], score_df["decisionScore"])
        }

        #3: Load TruMedia file and map decision scores
        trumedia_df = pd.read_csv(trumedia_file)
        trumedia_df["NormalizedName"] = trumedia_df["playerFullName"].str.strip()

        grades, decision_scores = [], []
        for name in trumedia_df["NormalizedName"]:
            match = score_map.get(name, {})
            grades.append(match.get("Grade", np.nan))
            decision_scores.append(match.get("decisionScore", np.nan))

        trumedia_df["Grade"] = grades
        trumedia_df["decisionScore"] = decision_scores

        #4: Insert new columns after playerFirstName and sort by decisionScore
        insert_index = list(trumedia_df.columns).index("playerFirstName") + 1
        cols = list(trumedia_df.columns)
        for col in ["Grade", "decisionScore", "NormalizedName"]:
            cols.remove(col)
        reordered = cols[:insert_index] + ["Grade", "decisionScore"] + cols[insert_index:]
        trumedia_final = trumedia_df[reordered].sort_values(by="decisionScore", ascending=False)

        #5: Export to Excel with color formatting
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            trumedia_final.to_excel(writer, index=False, sheet_name='TruMedia')
            sheet = writer.sheets['TruMedia']
            
            #Identify column indices
            score_col_idx = trumedia_final.columns.get_loc("decisionScore") + 1
            grade_col_idx = trumedia_final.columns.get_loc("Grade") + 1
            col_letter = get_column_letter(score_col_idx)
            score_range = f"{col_letter}2:{col_letter}{len(trumedia_final)+1}"

            #Conditional color scale for decisionScore
            rule = ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            )
            sheet.conditional_formatting.add(score_range, rule)

            #Gray out missing values
            gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            for row_idx, score in enumerate(trumedia_final["decisionScore"], start=2):
                if pd.isna(score):
                    sheet.cell(row=row_idx, column=score_col_idx).fill = gray_fill
                    sheet.cell(row=row_idx, column=grade_col_idx).fill = gray_fill

        output.seek(0)
        return send_file(
            io.BytesIO(output.read()),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='players_stats.xlsx'
        )

    except Exception as e:
        print("ERROR:", e)
        return f"An error occurred: {e}", 500

#Production Entrypoint
if __name__ == '__main__':
    #Run with Waitress server for production
    from waitress import serve
    serve(app, host='0.0.0.0', port=8080)
